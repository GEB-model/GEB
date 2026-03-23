"""Contains classes and methods for building the dependency tree of build methods, verification etc."""

import functools
import inspect
import logging
import tracemalloc
from datetime import datetime
from pathlib import Path
from time import time
from typing import Any, Iterable

import matplotlib.pyplot as plt
import networkx as nx

__all__: list[str] = ["build_method"]

from typing import Protocol

_BYTES_TO_MB: float = 1 / (1024 * 1024)


def validate_build_methods(
    tree: nx.DiGraph,
    methods: dict[str, Any],
    validate_order: bool = True,
    fix_order_if_broken: bool = False,
    logger: logging.Logger | None = None,
) -> dict[str, Any]:
    """Validate the methods in the dependency tree.

    Currently check the order of methods and whether they have the correct parameters.

    Args:
        tree: The dependency tree.
        methods: A dictionary of methods to validate.
        validate_order: If True, checks if methods depend on other methods that may come after them in the build file.
        fix_order_if_broken: If True, will attempt to fix the order of methods if validate_order fails. Only used if validate_order is True.
        logger: Logger to use for logging validation messages.

    Returns:
        The input methods if validation passes.

    Raises:
        ValueError: If a method depends on another method that is not a build function
        ValueError: If a method depends on another method that may come after it in the build file or not be present at all.
        ValueError: If a method has parameters that do not match the expected parameters.
    """
    for method in methods:
        if not tree.has_node(method):
            raise ValueError(
                f"Method {method} is not a build function.  If you are sure this should be build method, please decorate it with @build_method.'"
            )

        node = tree.nodes[method]
        if not node.get("attr", {}).get("_function_exists", False):
            raise ValueError(f"Method {method} is not a build function.")

    if validate_order:
        # Check if all dependencies are present in the requested methods.
        # This must be done regardless of fix_order_if_broken.
        for method in methods:
            # 1 is the method itself, 2 is the method + direct dependencies
            direct_dependencies = list(
                nx.dfs_postorder_nodes(tree.reverse(), method, depth_limit=2)
            )[:-1]
            for direct_dependency in direct_dependencies:
                if direct_dependency == method:
                    continue
                if direct_dependency not in methods:
                    if direct_dependency not in tree.nodes:
                        raise ValueError(
                            f"Method {method} depends on {direct_dependency}, "
                            "which is not a build function."
                        )
                    else:
                        raise ValueError(
                            f"Method {method} depends on {direct_dependency}, "
                            "which is missing from the requested methods."
                        )

        if fix_order_if_broken:
            # Check if the current order is already valid
            current_order_valid = True
            processed_in_check = set()
            for method in methods:
                # 1 is the method itself, 2 is the method + direct dependencies
                direct_deps = list(
                    nx.dfs_postorder_nodes(tree.reverse(), method, depth_limit=2)
                )[:-1]
                for dep in direct_deps:
                    if (
                        dep != method
                        and dep in methods
                        and dep not in processed_in_check
                    ):
                        current_order_valid = False
                        break
                if not current_order_valid:
                    break
                processed_in_check.add(method)

            if not current_order_valid:
                try:
                    # Find the minimal set of reorderings.
                    # We use a simple approach: find the first method that violates dependencies,
                    # and move its missing dependencies just before it.
                    # However, for a robust and deterministic "minimal" change,
                    # we can use the fact that the user likely wants to keep their existing order
                    # as much as possible.
                    # We'll use a stable topological sort that respects the original order
                    # where possible.
                    subgraph = tree.subgraph(methods.keys())
                    if not nx.is_directed_acyclic_graph(subgraph):
                        raise ValueError(
                            "Cannot fix order: cycle detected in requested methods."
                        )

                    changed_methods = set()

                    # To minimize changes, we can iterate and "pull up" dependencies.
                    new_order = list(methods.keys())
                    changed = True
                    while changed:
                        changed = False
                        for i in range(len(new_order)):
                            method = new_order[i]
                            deps = set(subgraph.predecessors(method))
                            if not deps:
                                continue

                            # Find the last position of any dependency in the current order
                            max_dep_idx = -1
                            for dep in deps:
                                dep_idx = new_order.index(dep)
                                if dep_idx > max_dep_idx:
                                    max_dep_idx = dep_idx

                            # If the last dependency is after the current method, move the method after it
                            if max_dep_idx > i:
                                method_to_move = new_order.pop(i)
                                new_order.insert(max_dep_idx, method_to_move)
                                changed = True
                                changed_methods.add(method_to_move)
                                break

                    methods = {node: methods[node] for node in new_order}

                    if logger is not None:
                        logger.warning(
                            f"The provided method order was invalid and has been auto-fixed. Moved methods: {changed_methods}"
                        )
                        logger.info(f"New build method order: {list(methods.keys())}")
                except nx.NetworkXUnfeasible:
                    raise ValueError("Cannot fix order: dependencies are circular.")
        else:
            processed_methods = set()
            for method in methods:
                # 1 is the method itself, 2 is the method + direct dependencies
                direct_dependencies = list(
                    nx.dfs_postorder_nodes(tree.reverse(), method, depth_limit=2)
                )[:-1]
                for direct_dependency in direct_dependencies:
                    if direct_dependency == method:
                        continue
                    if direct_dependency not in processed_methods:
                        raise ValueError(
                            f"Method {method} depends on {direct_dependency}, "
                            "which may come after this method in the build file or not be present at all."
                        )
                processed_methods.add(method)

    for method, args in methods.items():
        args_set = set(args) if args is not None else set()

        required_parameters = tree.nodes[method]["attr"]["_required_parameters"]
        if not set(required_parameters).issubset(args_set):
            raise ValueError(
                f"Method {method} has parameters {list(args.keys())}, "
                f"but expected parameters are {required_parameters}."
            )

        optional_parameters = tree.nodes[method]["attr"]["_optional_parameters"]
        if not set(args_set).issubset(set(required_parameters + optional_parameters)):
            raise ValueError(
                f"Method {method} has parameters {list(args.keys())}, "
                f"but required parameters are {required_parameters} and "
                f"optional parameters are {optional_parameters}."
            )

    return methods


class NamedCallable(Protocol):
    __name__: str

    def __call__(self, *args: Any, **kwargs: Any) -> Any: ...


class _build_method:
    def __init__(self, logger: logging.Logger | None = None) -> None:
        self.logger: logging.Logger | None = logger
        self.tree = nx.DiGraph()
        self.required_methods: set[str] = set()
        self.time_taken: dict[str, float] = {}
        self.peak_memory_usage: dict[str, int] = {}
        # Track what has already been flushed to the stats spreadsheet so
        # incremental calls to write_build_stats never produce duplicate rows.
        self._methods_written_to_stats: set[str] = set()
        self._disk_stats_written: bool = False

    def _resolve_logger(
        self, call_args: tuple[Any, ...] | None = None
    ) -> logging.Logger:
        """Resolve the logger for build method logging.

        Uses the logger on the bound model instance when available.
        For module-level operations, the logger must be assigned explicitly
        (for example via `build_method.logger = instance.logger`).

        Args:
            call_args: Positional arguments passed to a wrapped build method.

        Returns:
            Logger to use for logging build method messages.

        Raises:
            RuntimeError: If no logger is available.
        """
        if call_args:
            instance: Any = call_args[0]
            instance_logger: Any = getattr(instance, "logger", None)
            if isinstance(instance_logger, logging.Logger):
                return instance_logger
        if self.logger is None:
            raise RuntimeError(
                "build_method logger is not set. Ensure the owning class sets "
                "build_method.logger before invoking build methods."
            )
        return self.logger

    def __call__(
        self,
        required: bool,
        func: NamedCallable | None = None,
        depends_on: str | list[str] | None = None,
    ) -> NamedCallable:
        """Decorator to mark a method as a build method.

        Args:
            required: Whether the method is required to run.
            func: The function to decorate.
            depends_on: A method name or list of build_method that this method depends on.

        Returns:
            The decorated function.

        Raises:
            TypeError: if the decorator is used without parentheses.
        """

        def partial_decorator(func: NamedCallable) -> NamedCallable:
            @functools.wraps(func)
            def wrapper(*args: Any, **kwargs: Any) -> Any:
                active_logger: logging.Logger = self._resolve_logger(args)
                active_logger.info(f"Running method: {func.__name__}")
                for key, value in kwargs.items():
                    active_logger.debug(f"{func.__name__}.{key}: {value}")

                tracemalloc.start()
                start_time: float = time()
                value: Any = func(*args, **kwargs)
                end_time: float = time()
                _, peak = tracemalloc.get_traced_memory()
                tracemalloc.stop()

                elapsed_time: float = end_time - start_time

                self.time_taken[func.__name__] = elapsed_time
                self.peak_memory_usage[func.__name__] = peak
                active_logger.info(
                    f"Completed {func.__name__} in {elapsed_time:.2f} seconds with peak memory usage of {peak / 1024 / 1024:.2f} MB."
                )
                return value

            self.add_tree_node(func)

            if depends_on is not None:
                if isinstance(depends_on, str):
                    self.add_tree_edge(func, depends_on)
                elif isinstance(depends_on, list):
                    for item in depends_on:
                        self.add_tree_edge(func, item)
                else:
                    raise ValueError("depends_on must be a string or a list of strings")

            if required:
                self.required_methods.add(func.__name__)

            setattr(wrapper, "__is_build_method__", True)
            return wrapper

        if func is None:
            return partial_decorator
        else:
            raise TypeError(
                "Use @build_method() rather than @build_method without parentheses."
            )

    def add_tree_node(self, func: NamedCallable) -> None:
        """Add a node to the dependency tree."""
        parameters = inspect.signature(func).parameters
        required_parameters = [
            param.name
            for param in parameters.values()
            if param.default is inspect.Parameter.empty and param.name != "self"
        ]
        optional_parameters = [
            param.name
            for param in parameters.values()
            if param.default is not inspect.Parameter.empty
        ]
        self.tree.add_node(
            func.__name__,
            attr={
                "_function_exists": True,
                "_required_parameters": required_parameters,
                "_optional_parameters": optional_parameters,
            },
        )

    def add_tree_edge(self, func: NamedCallable, depends_on: str) -> None:
        """Add an edge to the dependency tree.

        Raises:
            ValueError: if a method depends on "setup_region" since everything depends on it.
                "setup_region" should therefore not be included in the dependency tree.

        """
        if depends_on == "setup_region":
            raise ValueError(
                "Everything depends on setup_region so we don't include it."
            )
        self.tree.add_edge(depends_on, func.__name__)

    def validate_tree(self) -> None:
        """Validate the dependency tree.

        Checks if all the node dependencies are present in the tree.

        Raises:
            ValueError: if a method depends on another method that is not a build function.
        """
        assert nx.is_directed_acyclic_graph(self.tree)
        for method in self.tree.nodes:
            depencencies = list(self.tree.predecessors(method))
            for dependency in depencencies:
                if (
                    not self.tree.nodes[dependency]
                    .get("attr", {})
                    .get("_function_exists", False)
                ):
                    raise ValueError(
                        f"Method {method} depends on {dependency}, "
                        "which is not a build function."
                    )
        self._resolve_logger().debug("Builder dependency tree validation passed.")

    def validate_methods(
        self,
        methods: dict[str, Any],
        validate_order: bool = True,
        fix_order_if_broken: bool = False,
    ) -> dict[str, Any]:
        """Validate the methods in the dependency tree.

        Currently check the order of methods and whether they have the correct parameters.

        Args:
            methods: A dictionary of methods to validate.
            validate_order: If True, checks if methods depend on other methods that may come after them in the build file.
            fix_order_if_broken: If True, will attempt to fix the order of methods if validate_order fails. Only used if validate_order is True.

        Returns:
            The input methods if validation passes.
        """
        active_logger: logging.Logger = self._resolve_logger()

        return validate_build_methods(
            self.tree,
            methods,
            validate_order=validate_order,
            fix_order_if_broken=fix_order_if_broken,
            logger=active_logger,
        )

    def export_tree(self) -> None:
        pos = nx.spring_layout(self.tree)
        nx.draw(self.tree, pos, with_labels=True, arrows=True)
        plt.savefig("dependency_graph.png")

    def get_dependencies(
        self, method: str, depth_limit: None | int = None
    ) -> list[str]:
        """Get all dependencies for a given method.

        Args:
            method: The method for which to find dependencies.
            depth_limit: Optional depth limit for the search.

        Returns:
            A list of methods that the specified method depends on.
        """
        # [:-1] is used to skip the method itself in the result
        return list(
            nx.dfs_postorder_nodes(self.tree.reverse(), method, depth_limit=depth_limit)
        )[:-1]

    def get_dependents(self, method: str, depth_limit: None | int = None) -> list[str]:
        """Get all methods that depend on a given method.

        Args:
            method: The method for which to find dependents.
            depth_limit: Optional depth limit for the search.

        Returns:
            A list of methods that depend on the specified method.
        """
        # [1:] is used to skip the method itself in the result
        return list(nx.dfs_preorder_nodes(self.tree, method, depth_limit=depth_limit))[
            1:
        ]

    def record_progress(self, progress_path: Path, method: str) -> None:
        """Record progress to txt progress file.

        Args:
            progress_path: Path to the progress file.
            method: Method that has been completed.
        """
        with open(progress_path, "a") as f:
            f.write(f"{method}\n")

    def read_progress(self, progress_path: Path) -> list[str]:
        """Get the list of methods that have been completed from the progress file.

        Args:
            progress_path: Path to the progress file.

        Returns:
            A list of methods that have been completed.

        Raises:
            ValueError: If duplicate methods are found in the progress file.
        """
        if not progress_path.exists():
            return []
        with open(progress_path, "r") as f:
            completed_methods: list[str] = f.read().splitlines()

        # Check for duplicates
        seen: set[str] = set()
        duplicates: list[str] = []
        for method in completed_methods:
            if method in seen:
                duplicates.append(method)
            seen.add(method)

        if duplicates:
            raise ValueError(
                f"Progress file corrupted with duplicates: {duplicates}. Possibly you run two concurrent builds at the same time? "
                f"Remove duplicates from the progress.txt and restart build with --continue."
            )

        return completed_methods

    def check_required_methods(self, methods: Iterable[str]) -> None:
        """Check that all required methods are present in the provided method list.

        Args:
            methods: A list of method names to check.

        Raises:
            ValueError: If any required method is missing.
        """
        missing_methods: set[str] = self.required_methods - set(methods)
        if missing_methods:
            raise ValueError(
                f"The following required methods are missing: {', '.join(missing_methods)}"
            )

    def log_statistics(self) -> None:
        """Log the time taken for each method in the dependency tree."""
        active_logger: logging.Logger = self._resolve_logger()
        total_time: float = sum(self.time_taken.values())

        sorted_by_time = sorted(
            self.time_taken.items(), key=lambda item: item[1], reverse=False
        )

        for method, time_taken in sorted_by_time:
            percentage: float = (time_taken / total_time) * 100
            active_logger.info(
                f"Method {method} took {time_taken:.2f} seconds ({percentage:.1f}%) and had peak memory usage of {self.peak_memory_usage[method] * _BYTES_TO_MB:.2f} MB."
            )

        sorted_by_memory = sorted(
            self.peak_memory_usage.items(), key=lambda item: item[1], reverse=False
        )

        for method, memory_usage in sorted_by_memory:
            percentage: float = (
                memory_usage / max(self.peak_memory_usage.values())
            ) * 100
            active_logger.info(
                f"Method {method} had peak memory usage of {memory_usage * _BYTES_TO_MB:.2f} MB ({percentage:.1f}%) and took {self.time_taken[method]:.2f} seconds."
            )

        active_logger.info(
            f"Total time taken: {total_time:.2f} seconds. Max memory usage: {max(self.peak_memory_usage.values()) * _BYTES_TO_MB:.2f} MB."
        )

    def write_build_stats(
        self,
        stats_path: Path,
        cluster_name: str,
        run_timestamp: datetime,
        cluster_dir: Path,
    ) -> None:
        """Append per-method memory/timing and per-folder disk-usage statistics to a shared Excel workbook.

        Writes two sheets:

        - ``memory_stats``: one row per build method with peak tracemalloc memory
          (in MB) and wall-clock time.
        - ``disk_stats``: one row per immediate subdirectory of ``cluster_dir``
          with its total recursive on-disk size.

        Rows are only appended for methods not yet recorded, so this can safely
        be called incrementally (e.g. after each method or at the end of a build)
        without producing duplicate rows.  The workbook is created on first use.

        Args:
            stats_path: Path to the Excel file (created if it does not exist).
            cluster_name: Short name of the cluster (e.g. ``"Europe_004"``).
            run_timestamp: Datetime at which the build run started.
            cluster_dir: Path to the scenario directory whose immediate
                subdirectories are measured for disk usage
                (e.g. ``large_scale6/Europe_004/base``).
        """
        import openpyxl  # noqa: PLC0415 – optional dependency, only needed here

        new_methods: dict[str, int] = {
            m: v
            for m, v in self.peak_memory_usage.items()
            if m not in self._methods_written_to_stats
        }
        if not new_methods and self._disk_stats_written:
            return

        stats_path = Path(stats_path)
        stats_path.parent.mkdir(parents=True, exist_ok=True)
        timestamp_str: str = run_timestamp.strftime("%Y-%m-%d %H:%M:%S")

        if stats_path.exists():
            wb = openpyxl.load_workbook(stats_path)
        else:
            wb = openpyxl.Workbook()
            wb.active.title = "memory_stats"
            wb.active.append(
                ["cluster", "run_started_at", "method", "peak_memory_mb", "elapsed_s"]
            )
            wb.create_sheet("disk_stats").append(
                ["cluster", "run_started_at", "folder", "size_mb"]
            )

        ws_mem = wb["memory_stats"]
        ws_disk = wb["disk_stats"]

        for method, peak_bytes in new_methods.items():
            ws_mem.append(
                [
                    cluster_name,
                    timestamp_str,
                    method,
                    round(peak_bytes * _BYTES_TO_MB, 1),
                    round(self.time_taken.get(method, float("nan")), 1),
                ]
            )

        # Write disk usage once per build run (snapshot at time of call).
        # Guarded by _disk_stats_written so incremental calls don't duplicate rows.
        if not self._disk_stats_written:
            cluster_dir = Path(cluster_dir)
            for folder in sorted(d for d in cluster_dir.iterdir() if d.is_dir()):
                folder_size_mb: float = round(
                    sum(f.stat().st_size for f in folder.rglob("*") if f.is_file())
                    * _BYTES_TO_MB,
                    1,
                )
                ws_disk.append(
                    [cluster_name, timestamp_str, folder.name, folder_size_mb]
                )
            self._disk_stats_written = True

        wb.save(stats_path)
        self._methods_written_to_stats.update(new_methods)

    @property
    def methods(self) -> list[str]:
        """Return the methods in the dependency tree, sorted alphabetically.

        Returns:
            A alphabetically sorted list of method names in the dependency tree.
        """
        return sorted(list(self.tree.nodes))


build_method = _build_method()
